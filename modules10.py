# Дополнительно устанавливаемые модули
# Модули документов
# PIP - Python Installation Package
# pip install openpyxl
# Чтение таблиц Excel
from openpyxl import load_workbook

# Загрузка рабочей книги (файл)
wb = load_workbook('excel/list.xlsx')
sheet = wb.active

# Читаем все строки
for item, price in sheet.iter_rows(values_only=True):
    print(f'Статья: {item} - Затраты: {price} руб.')
